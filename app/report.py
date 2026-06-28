from fileinput import filename
from app.logger import logger
from numpy import rint
from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime
import os


class ReportGenerator:

    def __init__(self):

        self.workbook = Workbook()

        self.sheet = self.workbook.active

        self.sheet.title = "Verification Result"

    def generate(self, records):

        headers = list(records[0].keys())

        # Write Headers
        for col, header in enumerate(headers, start=1):

            cell = self.sheet.cell(row=1, column=col)

            cell.value = header.replace("_", " ").title()

            cell.font = Font(bold=True)

        # Write Data
        for row_index, record in enumerate(records, start=2):

            for col_index, value in enumerate(record.values(), start=1):

                self.sheet.cell(
                    row=row_index,
                    column=col_index
                ).value = value

        # Leave two blank rows
        about_row = len(records) + 4

        self.sheet.cell(
            row=about_row,
            column=1
        ).value = "ABOUT"

        self.sheet.cell(
            row=about_row,
            column=1
        ).font = Font(
            bold=True,
            size=14
        )

        self.sheet.cell(
            row=about_row + 1,
            column=1
        ).value = "Software"

        self.sheet.cell(
            row=about_row + 1,
            column=2
        ).value = "Orunodoi Smart Verifier"

        self.sheet.cell(
            row=about_row + 2,
            column=1
        ).value = "Designed & Developed by"

        self.sheet.cell(
            row=about_row + 2,
            column=2
        ).value = "Ankur Dowarah"

        self.sheet.cell(
            row=about_row + 3,
            column=1
        ).value = "Version"

        self.sheet.cell(
            row=about_row + 3,
            column=2
        ).value = "1.0"

        self.sheet.cell(
            row=about_row + 4,
            column=1
        ).value = "Generated On"

        self.sheet.cell(
            row=about_row + 4,
            column=2
        ).value = datetime.now().strftime(
            "%d-%m-%Y %I:%M %p"
        )

        self.sheet.cell(
            row=about_row + 5,
            column=1
        ).value = "Copyright"

        self.sheet.cell(
            row=about_row + 5,
            column=2
        ).value = "© 2026 Ankur Dowarah. All Rights Reserved."

        # Create Output Folder
        os.makedirs("output", exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        filename = f"Orunodoi_Report_{timestamp}.xlsx"

        BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

        OUTPUT_DIR = os.path.join(BASE_DIR, "output")

        os.makedirs(OUTPUT_DIR, exist_ok=True)

        output_path = os.path.join(
    OUTPUT_DIR,
    filename
)

        self.workbook.save(output_path)

        print()

        logger.success("Output Saved Successfully")

        print(output_path)

        return output_path


       